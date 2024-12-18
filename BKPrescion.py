import pyvisa
import pandas as pd
import openpyxl
from datetime import datetime
import xlrd
import time

from openpyxl import load_workbook
from openpyxl import Workbook, load_workbook

import Conductivity7
import PH
import Temperature
from Graph import LivePlotter

ph = 0
is_running = True  # Shared flag variable

import pyvisa
import pandas as pd
from datetime import datetime
import time
from openpyxl import load_workbook

# Global variable for pH value
ph = 0
is_running = True  # Shared flag variable


def initialize_connection():
    # Create a PyVISA resource manager
    rm = pyvisa.ResourceManager()

    # Define the individual components of the resource address
    vid = '0xFFFF'
    pid = '0x9130'
    serial_number = '602360010736720014'
    device_address = f'USB0::0x{vid}::0x{pid}::{serial_number}::INSTR'

    # Open a connection to the power supply
    power_supply = rm.open_resource(device_address)

    # Return the power supply resource
    return power_supply


def startup(power_supply):
    # Query instrument identification
    power_supply.write('*IDN?')
    response = power_supply.read()
    print("Instrument identification:", response)

    # Query instrument firmware version
    power_supply.write(':SYSTem:VERSion?')
    response = power_supply.read()
    print("Firmware version:", response)

    # Additional startup tasks can be added here


def set_voltage(power_supply, voltage: float, channel: int):
    print("Sending " + str(voltage) + " Volts to channel " + str(channel))
    power_supply.write('INST CH' + str(channel))
    time.sleep(3)
    power_supply.write('VOLT ' + str(voltage))
    time.sleep(3)
    power_supply.close()


def set_current(power_supply, current: float, channel: int):
    time.sleep(3)
    print("Sending " + str(current) + " Amps to channel " + str(channel))
    current += .001  # .001 is needed because the power supply will be off by .001

    power_supply.write('INST CH' + str(channel))
    time.sleep(1.5)
    power_supply.write('CURRent' + str(current))
    time.sleep(1.5)
    power_supply.close()


def get_current_reading(power_supply, channel: int):
    power_supply.write('INST CH' + str(channel))
    current = power_supply.query(":MEAS:CURR?")
    power_supply.close()
    print('The current is: ' + str(current))
    return current


def get_voltage_reading(power_supply, channel: int):
    power_supply.write('INST CH' + str(channel))
    voltage = power_supply.query(":MEAS:VOLT?")
    power_supply.close()
    print('The Voltage is: ' + str(voltage))
    return voltage


def get_power_reading(instrument, channel):
    # Send the command to retrieve the power reading
    instrument.write('INST CH' + str(channel))
    power_reading = instrument.query('MEAS:POW?')

    # Return the power reading as a float value
    return float(power_reading)


def turn_off(power_supply):
    power_supply.write(":OUTP OFF")


def turn_on(power_supply):
    power_supply.write(":OUTP ON")


def save_variables_to_excel(file_path, channel):
    # You may include additional data retrieval here (e.g., temperature)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    voltage = get_voltage_reading(initialize_connection(), channel)
    current = get_current_reading(initialize_connection(), channel)
    power = get_power_reading(initialize_connection(), channel)
    temperature = (Temperature.get_temperature(8))
    conductivity = Conductivity.get_conductivity(7)
    global ph
    print('The ph is:' + str(ph))
    print('The voltage is: ' + str(voltage))
    print('The current is: ' + str(current))
    print('The power is: ' + str(power))
    print('The temperature is: ' + str(temperature))
    print('The Conductivity is: ' + str(conductivity))


    # Create a dictionary to store the data
    variables_dict = {
        'Time': [timestamp],
        'Voltage(V)': voltage,
        'Current(A)': current,
        'Power(W)': power,
        'PH': ph,
        'Temperature(C)': temperature,
        'Conductivity µS/cm 1.0': conductivity
    }

    # Convert the dictionary to a DataFrame
    df_new = pd.DataFrame(variables_dict)

    try:
        # Load an existing workbook or create a new one
        wb = load_workbook(file_path)
        sheet = wb.active
        next_row = sheet.max_row + 1

        # Append the new data to the existing sheet
        for index, row in df_new.iterrows():
            for col_num, value in enumerate(row, start=1):
                sheet.cell(row=next_row + index, column=col_num, value=value)

        # Save the changes to the workbook
        wb.save(file_path)
    except FileNotFoundError:
        # If the file doesn't exist, create a new one
        df_new.to_excel(file_path, index=False)

    print("Variables saved to Excel successfully.")


def monitor_variables(filepath, duration, channel):
    global ph

    count = 1
    plotter = LivePlotter()  # Assuming that `LivePlotter` is defined elsewhere
    x_value = 0
    # model = Machine_learning.perform_initial_training(filepath)
    while is_running:
        save_variables_to_excel(filepath, channel)
        # Code to be executed every `duration` minutes
        print("BK Precision channel " + str(channel) + ' has saved data ' + str(count) + " times")

        # Machine_learning.update_model_with_new_data(filepath, model) # Call the update method with the new data file
        plotter.update_plot(x_value, ph)  # updates graph
        x_value += 1

        time.sleep(duration)


"""""
while is_running is true. monitors ph and waits to activate devices when ph drops below ph_threshold 
param are ph threshold, stop ph, channel, and rpm
"""""


def active_devices(ph_threshold, wanted_ph, channel, rpm):
    ph_threshold = float(ph_threshold)
    wanted_ph = float(wanted_ph)
    while is_running:
        global ph
        ph = PH.get_ph(5)
        time.sleep(5)

        if ph <= ph_threshold:
            turn_on(initialize_connection())
            time.sleep(3)
            set_voltage(initialize_connection(), rpm, 3)

        elif ph >= wanted_ph:
            turn_off(initialize_connection())
